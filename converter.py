# -*- coding: utf-8 -*-
"""訂席資料表轉換器：人工維護的月份分頁 Excel → 機器友善的平面資料。

來源格式（115年訂席資料表）：
- 12 個月份分頁（1月～12月）
- row 1: 大標題、row 2: 區段標題、row 3: 欄位標題（★ 開頭為必填）、row 4: 圖例說明
- row 5 起為資料列，穿插「公休」列
"""
import re
from datetime import datetime, date, time

from openpyxl import load_workbook

MONTH_SHEET_PATTERN = re.compile(r"^\d{1,2}月$")
HEADER_ROW = 3   # 1-based：欄位標題列
DATA_START_ROW = 5

# 機器版總表的輸出欄位（轉換後的欄名 → 可接受的來源欄名清單，支援改名別名）
# 來源欄名以「移除 ★ 與空白後」比對，順序或位置變動不影響轉換
COLUMN_MAP = {
    "宴席日期": ["宴席日期"],
    "時段": ["時段"],
    "廳別": ["廳別"],
    "宴席名稱": ["宴席名稱"],
    "場控人員": ["場控人員"],
    "進館/會議時間": ["進館/會議時間"],
    "開席時間": ["開席時間"],
    "葷桌數": ["葷桌數"],
    "葷桌金額": ["葷桌金額"],
    "素食桌數": ["素食桌數"],
    "素食金額": ["素食金額", "素桌金額"],
    "素食套餐": ["素食套餐"],
    "素套金額": ["素套金額"],
    "預備桌": ["預備桌"],
    "主桌": ["主桌"],
    "主人數": ["主人數"],
    "訂金": ["訂金"],
    "菜單": ["菜單"],
    "確認菜單": ["確認菜單"],
    "試菜日期": ["試菜日期"],
    "服務費": ["服務費"],
    "婚禮佈置": ["婚禮佈置"],
    "主持/樂團": ["主持/樂團"],
    "聯絡人(主要)": ["聯絡人(主要)"],
    "關係": ["關係"],
    "電話(主要)": ["電話(主要)"],
    "聯絡人(次要)": ["聯絡人(次要)"],
    "電話(次要)": ["電話(次要)"],
    "承辦人": ["承辦人"],
    "訂席狀態": ["訂席狀態"],
    "最後聯繫日": ["最後聯繫日"],
    "備註": ["備註"],
}

# 這些欄位只有部分分頁有（新版才加的），缺少時不視為格式異常
OPTIONAL_COLUMNS = {"素食套餐", "素套金額"}

OUTPUT_HEADER = (
    ["來源分頁"]
    + list(COLUMN_MAP.keys())
    + ["電話(主要)正規化", "電話(次要)正規化"]
)


def _norm_header(value):
    if value is None:
        return ""
    return str(value).replace("★", "").replace(" ", "").strip()


def normalize_phone(value):
    """電話正規化：只留數字，供比對用（0988-602-016 → 0988602016）。"""
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits


def _cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def parse_workbook(path_or_buffer):
    """解析整本訂席資料表。

    回傳 (records, issues)：
    - records: list[dict]，每筆一場宴席（含公休列，訂席狀態欄標為「公休」）
    - issues: list[str]，格式異常說明（給轉換報告用）
    """
    wb = load_workbook(path_or_buffer, data_only=True, read_only=True)
    records = []
    issues = []

    for sheet_name in wb.sheetnames:
        if not MONTH_SHEET_PATTERN.match(sheet_name):
            # 非月份分頁（如「除夕」外帶訂單、說明頁）不屬於宴席檔期，
            # 安靜略過、不列入異常通知
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < HEADER_ROW:
            issues.append(f"{sheet_name}：列數不足，略過")
            continue

        headers = [_norm_header(v) for v in rows[HEADER_ROW - 1]]
        col_index = {}
        for out_name, aliases in COLUMN_MAP.items():
            for alias in aliases:
                key = _norm_header(alias)
                if key in headers:
                    col_index[out_name] = headers.index(key)
                    break
        if "宴席日期" not in col_index:
            issues.append(f"{sheet_name}：找不到「宴席日期」欄，略過整頁")
            continue
        missing = [
            n for n in COLUMN_MAP
            if n not in col_index and n not in OPTIONAL_COLUMNS
        ]
        if missing:
            issues.append(f"{sheet_name}：缺少欄位 {missing}，該等欄位輸出為空白")

        for row_no, row in enumerate(rows[DATA_START_ROW - 1:], start=DATA_START_ROW):
            raw_date = row[col_index["宴席日期"]] if col_index["宴席日期"] < len(row) else None
            if raw_date is None or str(raw_date).strip() == "":
                continue  # 空列
            if not isinstance(raw_date, (datetime, date)):
                issues.append(
                    f"{sheet_name} 第{row_no}列：宴席日期「{raw_date}」不是日期格式，略過"
                )
                continue

            record = {"來源分頁": sheet_name}
            for out_name in COLUMN_MAP:
                idx = col_index.get(out_name)
                value = row[idx] if idx is not None and idx < len(row) else None
                record[out_name] = _cell_to_str(value)

            # 公休列：時段欄為「公休」
            if record.get("時段") == "公休" and not record.get("訂席狀態"):
                record["訂席狀態"] = "公休"

            record["電話(主要)正規化"] = normalize_phone(record.get("電話(主要)"))
            record["電話(次要)正規化"] = normalize_phone(record.get("電話(次要)"))
            records.append(record)

    wb.close()
    records.sort(key=lambda r: (r["宴席日期"], r["時段"], r["廳別"]))
    return records, issues


def records_to_rows(records):
    """轉成寫入 Google Sheets 用的二維陣列（含標題列）。"""
    rows = [OUTPUT_HEADER]
    for r in records:
        rows.append([r.get(col, "") for col in OUTPUT_HEADER])
    return rows


if __name__ == "__main__":
    import sys

    sys.stdout.reconfigure(encoding="utf-8")
    src = sys.argv[1] if len(sys.argv) > 1 else (
        r"C:\Users\鴻逵\Downloads\維多利亞宴會館_115年_訂席資料表_完整版.xlsx"
    )
    records, issues = parse_workbook(src)
    print(f"共解析 {len(records)} 筆")
    closed = sum(1 for r in records if r["訂席狀態"] == "公休")
    print(f"其中公休 {closed} 筆、宴席 {len(records) - closed} 筆")
    print("\n--- 格式異常 ---")
    for issue in issues or ["（無）"]:
        print(" -", issue)
    print("\n--- 前 5 筆預覽 ---")
    for r in records[:5]:
        print(
            f"{r['宴席日期']} {r['時段']:<3} {r['廳別']:<10} {r['宴席名稱'][:20]:<20} "
            f"電話={r['電話(主要)正規化'] or '-'} 狀態={r['訂席狀態'] or '-'}"
        )
